"""
multitester.py
Run all 4 stylized fact tests across multiple tickers and timeframes,
then export results to Excel and a combined multi-ticker plot.

Edit TICKERS / INTERVALS / START_DATE / END_DATE at the top, then run:
    python multitester.py
"""
import io
import sys
import time
import logging
import traceback

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from utilities.data import DataManager
from facts_test.volclustering import VolatilityClustering
from facts_test.decay import SlowDecay
from facts_test.leverage import LeverageEffect
from facts_test.volvolcorr import VolVolCorr

# ── USER CONFIGURATION ─────────────────────────────────────────────────────────
TICKERS = [
    "NVDA", "META", "AMD",         # Tech, Social & Semiconductors
    "HD", "SBUX", "SPY", "QQQ",           # Consumer Discretionary & Dining
     "PFE",           # Healthcare & Pharmaceuticals
    "XOM", "CVX",                  # Energy & Oil Heavyweights
    "V",                     # Global Payment Processors
     "BA",                   # Industrials & Aerospace
    "DIS",                  # Media & Entertainment
    "KO", "PG", "VNQ"              # Consumer Staples & Real Estate ETF
]
INTERVALS   = ["1m"]
START_DATE  = "2021-01-01"
END_DATE    = "2026-01-01"
OUTPUT_FILE = "multitester_results.xlsx"
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.WARNING, format='%(levelname)s: %(message)s')

# ── INTERVAL-SPECIFIC PARAMETERS ──────────────────────────────────────────────
_SD_MAX_LAGS  = {'1m': 2000, '5m': 1000, '1h': 500,  '1d': 200}
_LE_MAX_LAGS  = {'1m': 180,  '5m': 60,   '1h': 50,   '1d': 30}
_VVC_MAX_LAGS = {'1m': 500,  '5m': 300,  '1h': 150,  '1d': 150}
# ──────────────────────────────────────────────────────────────────────────────


class _Quiet:
    """Context manager: suppress stdout and stderr (prints inside testers)."""
    def __enter__(self):
        self._out, self._err = sys.stdout, sys.stderr
        sys.stdout = sys.stderr = io.StringIO()

    def __exit__(self, *_):
        sys.stdout, sys.stderr = self._out, self._err


def _safe(fn, fallback=('INCONCLUSIVE', None)):
    """Run fn(); on any exception return fallback."""
    try:
        return fn()
    except Exception:
        return fallback


# ── INDIVIDUAL FACT FORMATTERS ─────────────────────────────────────────────────

def fact1(returns, ticker):
    """Returns (text, plot_data) for Volatility Clustering."""
    with _Quiet():
        t = VolatilityClustering(returns, ticker)
        c1, sig = t.compute_c1(max_lag=40, plot=False)
    text = f'CONFIRMED ({len(sig)} lags)' if len(sig) > 0 else 'NOT DETECTED'
    pdata = None
    if len(c1) > 1:
        pdata = {'lags': np.arange(1, len(c1)).tolist(), 'c1': c1[1:].tolist()}
    return text, pdata


def fact2(returns, ticker, interval):
    """Returns (text, plot_data) for Slow Decay."""
    ml = _SD_MAX_LAGS.get(interval, 500)
    with _Quiet():
        t = SlowDecay(returns, ticker)
        r = t.compute_decay(max_lag=ml, plot=False)
    if r is None:
        return 'INCONCLUSIVE', None
    if r['slow_decay_confirmed']:
        b1, b2 = r['beta_alpha1'], r['beta_alpha2']
        b = b1 if (b1 is not None and 0 < b1 < 1) else b2
        text = f"CONFIRMED (Beta={b:.3f})"
    else:
        b1, b2 = r['beta_alpha1'], r['beta_alpha2']
        b = b1 if b1 is not None else b2
        b_str = f'{b:.3f}' if b is not None else 'N/A'
        text = f"NOT CONFIRMED (Beta={b_str})"
    pdata = {
        'lags':  r['lags_alpha1'],
        'acf':   r['acf_alpha1'],
        'beta':  r['beta_alpha1'],
        'A':     r['A_alpha1'],
    }
    return text, pdata


def fact3(returns, ticker, interval):
    """Returns (text, plot_data) for Leverage Effect."""
    ml = _LE_MAX_LAGS.get(interval, 50)
    with _Quiet():
        t = LeverageEffect(returns, ticker)
        r = t.compute_leverage(max_lag=ml, plot=False)
    if r is None:
        return 'INCONCLUSIVE', None
    if r['leverage_detected']:
        text = f"CONFIRMED (L={r['min_L']:.2f})"
    else:
        min_l = r['min_L']
        l_str = f"{min_l:.2f}" if np.isfinite(min_l) else 'N/A'
        text = f"NOT DETECTED (min L={l_str})"
    pdata = {'lags': r['lags'], 'L': r['L_values'], 'null_lower': r['null_lower']}
    return text, pdata


def fact4(returns, volume, ticker, interval):
    """Returns (text, plot_data) for Vol-Vol Correlation."""
    ml = _VVC_MAX_LAGS.get(interval, 100)
    with _Quiet():
        t = VolVolCorr(returns, volume, ticker)
        r = t.compute_correlation(max_lag=ml, plot=False)
    if r is None:
        return 'INCONCLUSIVE', None
    lags = np.array(r['lags'])
    c0 = float(np.array(r['corr_abs'])[lags == 0][0])
    text = f"CONFIRMED (C(0)={c0:.3f})" if r['corr_confirmed'] else f"NOT DETECTED (C(0)={c0:.3f})"
    pdata = {'lags': r['lags'], 'corr_abs': r['corr_abs'], 'null_upper': r['null_upper']}
    return text, pdata


# ── MAIN RUNNER ───────────────────────────────────────────────────────────────

def run_single(ticker, interval, dm):
    """Fetch data and run all 4 facts. Returns (results dict, plot_data dict)."""
    df, returns, report = dm.get_data(ticker, START_DATE, END_DATE, interval)
    if df.empty:
        print(f"    NO DATA — {report}")
        return {f'f{i}': 'NO DATA' for i in range(1, 5)}, {}

    print(f"    {len(returns)} bars | {report}")
    results = {}
    plot_data = {}
    volume = df['Volume']

    facts = [
        ('f1', lambda: fact1(returns, ticker)),
        ('f2', lambda: fact2(returns, ticker, interval)),
        ('f3', lambda: fact3(returns, ticker, interval)),
        ('f4', lambda: fact4(returns, volume, ticker, interval)),
    ]

    for key, fn in facts:
        t0 = time.time()
        result = _safe(fn)
        elapsed = time.time() - t0
        if isinstance(result, tuple):
            val, pdata = result
        else:
            val, pdata = result, None
        tag = '✅' if val.startswith(('CONFIRMED', 'PASSED')) else ('⚠️' if 'INCONCLUSIVE' in val or 'NO DATA' in val else '❌')
        print(f"    {key.upper():<4} {tag}  {val}  ({elapsed:.1f}s)")
        results[key] = val
        if pdata is not None:
            plot_data[key] = pdata

    return results, plot_data


# ── COMBINED PLOT ──────────────────────────────────────────────────────────────

def _plot_combined(ticker_plot_data, interval, output_file):
    """
    4-subplot figure: one panel per stylized fact, each with all tickers
    overlaid as colored dots+lines and a shared legend.

    F2 uses log-log axes to highlight the power-law decay slope.
    """
    valid = [(t, p) for t, p in ticker_plot_data if p]
    if not valid:
        print("  No plot data collected — skipping combined plot.")
        return

    n = len(valid)
    cmap = plt.cm.tab20
    colors = {t: cmap(i % 20) for i, (t, _) in enumerate(valid)}

    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle(
        f'Stylized Facts — Multi-Ticker  |  {interval}  |  {START_DATE} → {END_DATE}',
        fontsize=13, fontweight='bold',
    )
    ax1, ax2 = axes[0, 0], axes[0, 1]
    ax3, ax4 = axes[1, 0], axes[1, 1]

    null_ci1_vals, null_lower3_vals, null_upper4_vals = [], [], []

    for ticker, pdata in valid:
        c = colors[ticker]
        kw = dict(color=c, linewidth=0.5, alpha=0.8,
                  marker='o', markersize=3, label=ticker)

        # F1: Volatility Clustering — ACF(|r|) vs lag
        d1 = pdata.get('f1')
        if d1:
            ax1.plot(d1['lags'], d1['c1'], **kw)

        # F2: Slow Decay — log-log ACF(|r|) vs lag (significant range only)
        d2 = pdata.get('f2')
        if d2 and len(d2['lags']) > 1:
            lags2 = np.array(d2['lags'])
            acf2  = np.array(d2['acf'])
            pos   = (lags2 > 0) & (acf2 > 0)
            if pos.sum() > 1:
                ax2.plot(np.log(lags2[pos]), np.log(acf2[pos]), **kw)
                # Overlay power-law fit line (no label to avoid legend clutter)
                beta, A = d2.get('beta'), d2.get('A')
                if beta is not None and A is not None and 0 < beta < 1:
                    lx = np.log(lags2[pos])
                    ax2.plot(lx, np.log(A) - beta * lx,
                             color=c, linewidth=1.0, linestyle='--', alpha=0.5)

        # F3: Leverage Effect — L(tau) vs tau
        d3 = pdata.get('f3')
        if d3:
            L    = np.array(d3['L'])
            lags3 = np.array(d3['lags'])
            fin  = np.isfinite(L)
            ax3.plot(lags3[fin], L[fin], **kw)
            null_lower3_vals.append(d3['null_lower'])

        # F4: Vol-Vol Correlation — C(tau) vs tau
        d4 = pdata.get('f4')
        if d4:
            ax4.plot(d4['lags'], d4['corr_abs'], **kw)
            null_upper4_vals.append(d4['null_upper'])

    # ── Reference lines ────────────────────────────────────────────────────────
    ax1.axhline(0, color='gray', linestyle='--', linewidth=0.7, zorder=0)
    ax1.set_title('F1: Volatility Clustering  —  ACF(|r|)')
    ax1.set_xlabel('Lag τ')
    ax1.set_ylabel('C₁(τ)')

    ax2.axhline(0, color='gray', linestyle='--', linewidth=0.7, zorder=0)
    ax2.set_title('F2: Slow Decay  —  log-log ACF(|r|)  (dashed = power-law fit)')
    ax2.set_xlabel('ln(τ)')
    ax2.set_ylabel('ln(ACF)')

    if null_lower3_vals:
        avg3 = float(np.mean(null_lower3_vals))
        ax3.axhline(avg3, color='red', linestyle='--', linewidth=0.8,
                    label=f'avg null lower ({avg3:.3f})', zorder=0)
    ax3.axhline(0, color='gray', linestyle='--', linewidth=0.7, zorder=0)
    ax3.axvline(0, color='gray', linestyle='--', linewidth=0.7, zorder=0)
    ax3.set_title('F3: Leverage Effect  —  L(τ) = Corr(r_t, r²_{t+τ})')
    ax3.set_xlabel('Lag τ')
    ax3.set_ylabel('L(τ)')

    if null_upper4_vals:
        avg4 = float(np.mean(null_upper4_vals))
        ax4.axhline(avg4, color='red', linestyle='--', linewidth=0.8,
                    label=f'avg null upper ({avg4:.3f})', zorder=0)
    ax4.axhline(0, color='gray', linestyle='--', linewidth=0.7, zorder=0)
    ax4.axvline(0, color='gray', linestyle='--', linewidth=0.7, zorder=0)
    ax4.set_title('F4: Vol-Vol Correlation  —  Corr(V_t, |r_{t+τ}|)')
    ax4.set_xlabel('Lag τ')
    ax4.set_ylabel('C(τ)')

    # ── Single shared legend at the bottom ────────────────────────────────────
    seen, handles, labels = set(), [], []
    for ax in (ax1, ax2, ax3, ax4):
        for h, lbl in zip(*ax.get_legend_handles_labels()):
            if lbl not in seen and not lbl.startswith('avg'):
                handles.append(h)
                labels.append(lbl)
                seen.add(lbl)
    # Red reference lines as a single legend entry
    for ax in (ax3, ax4):
        for h, lbl in zip(*ax.get_legend_handles_labels()):
            if lbl.startswith('avg') and lbl not in seen:
                handles.append(h)
                labels.append(lbl)
                seen.add(lbl)

    fig.legend(handles, labels,
               loc='lower center', ncol=min(n + 2, 8),
               fontsize=8, bbox_to_anchor=(0.5, 0.0))

    plt.tight_layout(rect=[0, 0.06, 1, 1])

    plot_file = output_file.replace('.xlsx', f'_{interval}_plot.png')
    plt.savefig(plot_file, dpi=150, bbox_inches='tight')
    print(f'\n📊 Combined plot saved to: {plot_file}')
    plt.show()


# ── EXCEL WRITER ──────────────────────────────────────────────────────────────

HEADERS = [
    'Ticker', 'Time',
    'F1: Vol. Clustering', 'F2: Slow Decay',
    'F3: Leverage Effect', 'F4: Vol-Vol Corr.',
]

_GREEN  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
_RED    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
_ALT    = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


def _fill_for(value: str):
    v = (value or '').upper()
    if v.startswith(('CONFIRMED', 'PASSED')):
        return _GREEN
    if v.startswith(('FAILED', 'NOT CONFIRMED', 'NOT DETECTED', 'NOT SIGNIFICANT')):
        return _RED
    if v.startswith(('INCONCLUSIVE', 'NO DATA', 'ERROR')):
        return _YELLOW
    return None


def _write_data_rows(ws, rows, start_row):
    """Write data rows into ws starting at start_row, with colour formatting."""
    existing_tickers = []
    for r in range(3, start_row):
        val = ws.cell(r, 1).value
        if val and val not in existing_tickers:
            existing_tickers.append(val)

    new_tickers = list(dict.fromkeys(t for t, _, _ in rows))
    all_tickers = existing_tickers + [t for t in new_tickers if t not in existing_tickers]

    for row_idx, (ticker, interval, facts) in enumerate(rows, start=start_row):
        alt = all_tickers.index(ticker) % 2 == 1

        def cell(col, val=''):
            c = ws.cell(row_idx, col, val)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            return c

        cell(1, ticker)
        cell(2, interval)

        if alt:
            ws.cell(row_idx, 1).fill = _ALT
            ws.cell(row_idx, 2).fill = _ALT

        for col_offset, key in enumerate(['f1', 'f2', 'f3', 'f4'], start=3):
            val = facts.get(key, '')
            c = cell(col_offset, val)
            fill = _fill_for(val)
            c.fill = fill if fill else (_ALT if alt else PatternFill())


def write_excel(rows, output_file):
    import os

    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb['Stylized Facts'] if 'Stylized Facts' in wb.sheetnames else wb.active
        start_row = ws.max_row + 1
        print(f'  Appending {len(rows)} rows after row {start_row - 1} in existing file.')
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Stylized Facts'

        note_cell = ws.cell(1, 1, f'Dates: {START_DATE} if possible till {END_DATE}')
        note_cell.font = Font(italic=True, color='808080')

        hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF')
        center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col, h in enumerate(HEADERS, start=1):
            c = ws.cell(2, col, h)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = center

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 6
        for col in range(3, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = 'A3'

        start_row = 3
        print(f'  Creating new file: {output_file}')

    _write_data_rows(ws, rows, start_row)
    wb.save(output_file)
    print(f'\n✅ Results saved to: {output_file}')


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

def main():
    dm = DataManager()
    rows = []
    all_plot_data = []   # list of (ticker, interval, plot_data_dict)
    total = len(TICKERS) * len(INTERVALS)
    done  = 0

    t_start = time.time()

    for ticker in TICKERS:
        for interval in INTERVALS:
            done += 1
            print(f'\n[{done}/{total}] {ticker} @ {interval}')
            try:
                facts, pdata = run_single(ticker, interval, dm)
            except Exception:
                print(f'  ❌ Unhandled exception:\n{traceback.format_exc()}')
                facts = {f'f{i}': 'ERROR' for i in range(1, 5)}
                pdata = {}
            rows.append((ticker, interval, facts))
            all_plot_data.append((ticker, interval, pdata))

    elapsed = time.time() - t_start
    print(f'\nTotal runtime: {elapsed/60:.1f} min')
    write_excel(rows, OUTPUT_FILE)

    for interval in INTERVALS:
        interval_data = [(t, p) for t, iv, p in all_plot_data if iv == interval]
        _plot_combined(interval_data, interval, OUTPUT_FILE)


if __name__ == '__main__':
    main()
